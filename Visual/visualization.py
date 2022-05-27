import matplotlib.pyplot as plt
import numpy as np
import json
from typing import Union
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

alignment_center = Alignment(horizontal='center', vertical='center')
def Draw(data:np.ndarray,labels:dict,pic_path:str)->None:
    fig,ax = plt.subplots(nrows=1,ncols=len(labels),figsize=(2*len(labels),6))
    colors = ["#FF0000","#00FF00","#0000FF","#FFFF00","#00FFFF","#FF00FF","#FFF000","#00FFF0","#F000FF","#F00000","#00F000","#0000F0","#F00F00","#000000"]
    for i,j in enumerate(labels.items()):
        model_name, res_label = j
        # se存储每个聚类的样本个数
        se =  {str(i):0 for i in set(res_label)}
        #used_colors = list()
        for k in se:
            num = int(k)
            se[k]=str((res_label==num).sum())
            pdata = data[res_label==num]
            color = colors[num]
            ax[i].set_title(model_name)
            ax[i].scatter(pdata[:,0],pdata[:,1],s=5,color=color)
            #used_colors.append(num)
        #KMeans确实是n_clusters类,DBSCAN得看参数
        #ax[i].legend(used_colors)
        with open(f"{model_name}.json","w",encoding="utf-8") as f:
            json.dump(se,f)
    plt.savefig(pic_path)

def make_xlsx_header(ws:Worksheet,labels:dict,indexs:list,column:list)->None:
    size = len(indexs)
    for i,j in enumerate(list(labels.keys())):
        ws.merge_cells(start_row=i*size+2,end_row=(i+1)*size+1,start_column=1,end_column=1)
        ws.cell(row=i*size+2,column=1).value=j
        ws.cell(row=i*size+2,column=1).alignment = alignment_center
        for s,t in enumerate(indexs):
            ws.cell(row=i*size+2+s,column=2).value=t
            ws.cell(row=i*size+2+s,column=2).alignment = alignment_center
    ws.cell(row=1,column=2).value = "n_components"
    ws.cell(row=1,column=2).alignment = alignment_center
    for j in column:
        ws.cell(row=1,column=j+1).value=j
        ws.cell(row=1,column=j+1).alignment = alignment_center

def make_xlsx_body(ws:Worksheet,row:int,col:int,value:str)->None:
    ws.cell(row,col).value=value
    ws.cell(row,col).alignment = alignment_center
