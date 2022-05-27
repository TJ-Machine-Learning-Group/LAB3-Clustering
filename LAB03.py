from random import randint
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from sklearn.cluster import DBSCAN
from ClusterValidityIndex.CVI import Eval, DBI, CH, Silhouette
from DataPrepare.dataAnalyzeandPrepare import GetData, GetNorData, GetPCAData
import json
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment

def Predict(models: dict, data: np.ndarray) -> dict:
    labels = dict()
    #Kmeans: core and distance
    labels['kmeans'] = models['kmeans'].fit_predict(data)
    #DBSCN: near neighbor
    labels['dbscan'] = models['dbscan'].fit_predict(data)
    return labels


if __name__ == "__main__":
    models = dict()
    #初始化模型，要调参可以在这里调
    random_state = randint(1, 100)  #给定一个随机种子（划分数据集、确定初始点等）
    n_clusters = 12
    models['kmeans'] = KMeans(n_clusters=n_clusters, random_state=random_state)

    eps = 0.5
    min_samples = 10
    models['dbscan'] = DBSCAN(eps=eps, min_samples=min_samples)

    #标准化+PCA降维 保留2个分量
    csv_path = "./Live_20210128.csv"
    data = GetPCAData(csv_path=csv_path, n_components=2)
    labels = Predict(models, data)  #data是(m,6)的数组，默认使用欧氏距离（要用其它距离可以先算出来

    fig,ax = plt.subplots(nrows=1,ncols=2,figsize=(8, 6))
    colors = ["#FF0000","#00FF00","#0000FF","#FFFF00","#00FFFF","#FF00FF","#FFF000","#00FFF0","#F000FF","#F00000","#00F000","#0000F0","#F00F00","#000000"]
    for i,j in enumerate(labels.items()):
        model_name, res_label = j
        # se存储每个聚类的样本个数
        se =  {str(i):0 for i in set(res_label)}
        used_colors = list()
        #print(se)
        for k in se:
            num = int(k)
            se[k]=str((res_label==num).sum())
            pdata = data[res_label==num]
            color = colors[num]
            ax[i].set_title(model_name)
            ax[i].scatter(pdata[:,0],pdata[:,1],s=5,color=color)
            used_colors.append(num)
        #KMeans确实是n_clusters类,DBSCAN得看参数
        ax[i].legend(used_colors)
        with open(f"{model_name}.json","w",encoding="utf-8") as f:
            json.dump(se,f)
        eval_data = Eval(data, res_label)
    plt.savefig("test.jpg")
    indexs = ["number of clusters","num of noise points","Silhouette Coefficient","Calinski-Harabaz Index","Davies-Bouldin Index","all"]
    size = len(indexs)
    wb = Workbook()
    ws = wb.active
    alignment_center = Alignment(horizontal='center', vertical='center')
    for i,j in enumerate(list(labels.keys())):
        ws.merge_cells(start_row=i*size+2,end_row=(i+1)*size+1,start_column=1,end_column=1)
        ws.cell(row=i*size+2,column=1).value=j
        ws.cell(row=i*size+2,column=1).alignment = alignment_center
        for s,t in enumerate(indexs):
            ws.cell(row=i*size+2+s,column=2).value=t
            ws.cell(row=i*size+2+s,column=2).alignment = alignment_center
    for j in range(2,9):
        ws.cell(row=1,column=j+1).value=j
        ws.cell(row=1,column=j+1).alignment = alignment_center
        data=GetPCAData(csv_path=csv_path,n_components=j)
        for i,k in enumerate(labels.items()):
            model_name, res_label = k
            eval_data = Eval(data, res_label)
            for s in range(len(eval_data)):
                ws.cell(row=i*size+2+s,column=j+1).value=eval_data[s]
                ws.cell(row=i*size+2+s,column=j+1).alignment = alignment_center
    wb.save('test.xlsx')
